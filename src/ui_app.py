"""
Aspen Plus 自动调参 UI 工具
支持自动开启 Aspen Plus、浏览模块/流股变量、配置输入范围、
选择输出变量、批量驱动 Aspen 模拟、结果保存到 Excel

Copyright (c) 2026 lijunsen & DeepSeek
"""

import os
import sys
import csv
import io
import time
import json
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
from itertools import product

# ── 基础路径（PyInstaller 兼容） ──
def _get_exe_dir():
    """返回 exe 所在目录（源码运行则返回项目根目录）"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── 全局未捕获异常处理 ──
_CRASH_LOG = os.path.join(_get_exe_dir(), "crash_log.txt")

def _global_excepthook(exc_type, exc_value, exc_tb):
    msg = "=" * 60 + "\n"
    msg += f"时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
    msg += f"类型: {exc_type.__name__}\n"
    msg += f"内容: {exc_value}\n"
    msg += "跟踪:\n" + "".join(traceback.format_exception(exc_type, exc_value, exc_tb)) + "\n"
    msg += "=" * 60 + "\n"
    try:
        with open(_CRASH_LOG, "a", encoding="utf-8") as f:
            f.write(msg)
        try:
            import tkinter.messagebox as _mb
            _mb.showerror("程序出错",
                          f"{exc_type.__name__}: {exc_value}\n\n"
                          f"详情已保存至:\n{_CRASH_LOG}")
        except:
            pass
    except:
        pass
    sys.__excepthook__(exc_type, exc_value, exc_tb)

sys.excepthook = _global_excepthook

def _save_crash(msg):
    """将崩溃信息写入 crash_log.txt"""
    full = "=" * 60 + "\n"
    full += f"时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n{msg}\n"
    full += "=" * 60 + "\n"
    try:
        with open(_CRASH_LOG, "a", encoding="utf-8") as f:
            f.write(full)
    except:
        pass

# ── PyInstaller 打包兼容 ──
# 运行时：sys._MEIPASS 指向解压临时目录
# 源码时：__file__ 指向当前目录
if getattr(sys, 'frozen', False):
    _base_path = sys._MEIPASS
else:
    _base_path = os.path.dirname(os.path.abspath(__file__))

if _base_path not in sys.path:
    sys.path.insert(0, _base_path)

from aspen_interface import AspenInterface
from aspen_interface import log as _aspen_log

# 尝试导入 COM 初始化（可能被 PyInstaller 排除）
try:
    import pythoncom
    _HAVE_PYTHONCOM = True
except ImportError:
    _HAVE_PYTHONCOM = False


class AspenUI:
    """Aspen Plus 自动化调参图形界面"""

    def __init__(self, root):
        self.root = root
        self.root.title("Aspen Plus 自动调参工具  —  © 2026 lijunsen & DeepSeek")
        self.root.geometry("1280x800")
        self.root.minsize(1024, 680)

        # 核心实例
        self.aspen = AspenInterface()
        self.aspen_connected = False

        # 设置窗口图标
        self._set_window_icon()

        # 数据缓存
        self.blocks_info = []
        self.streams_info = []
        self.all_input_vars = {}
        self.all_output_vars = {}

        # 用户配置
        self.input_configs = []
        self.output_configs = []

        # 结果存储
        self.simulation_results = []

        # 样式
        self.style = ttk.Style()
        self.style.theme_use("vista")
        self._setup_styles()

        # 构建界面
        self._build_ui()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        # 防止 tkinter 吞掉按钮回调异常
        self.root.report_callback_exception = self._tk_callback_exception

    def _tk_callback_exception(self, exc_type, exc_value, exc_tb):
        """捕获 tkinter 事件回调中抛出的异常（默认会被 tkinter 静默吞掉）"""
        msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        _save_crash(f"[tkinter 回调异常] {msg}")
        try:
            messagebox.showerror("操作出错",
                f"{exc_type.__name__}: {exc_value}\n\n详情已保存至 crash_log.txt")
        except:
            pass

    def _set_window_icon(self):
        """设置窗口和任务栏图标（源码/打包模式自动适配路径）"""
        try:
            if getattr(sys, 'frozen', False):
                ico_dir = os.path.join(sys._MEIPASS, 'ico')
            else:
                ico_dir = os.path.join(_get_exe_dir(), 'ico')
            ico_path = os.path.join(ico_dir, '128x128.ico')
            if os.path.isfile(ico_path):
                self.root.iconbitmap(default=ico_path)
        except:
            pass

    def _setup_styles(self):
        """配置 ttk 样式——专业蓝白配色"""
        bg = "#F5F7FA"
        self.root.configure(bg=bg)

        self.style.configure(".", font=("微软雅黑", 10), background=bg)
        self.style.configure("TFrame", background=bg)
        self.style.configure("TLabel", background=bg, font=("微软雅黑", 10))
        self.style.configure("TLabelframe", background=bg, font=("微软雅黑", 10, "bold"))
        self.style.configure("TLabelframe.Label", background=bg, font=("微软雅黑", 10, "bold"))

        # 标题
        self.style.configure("Title.TLabel", font=("微软雅黑", 14, "bold"), foreground="#1E3A5F")
        self.style.configure("Header.TLabel", font=("微软雅黑", 11, "bold"), foreground="#334155")
        self.style.configure("Status.TLabel", font=("微软雅黑", 9), foreground="#64748B")

        # 按钮
        self.style.configure("Action.TButton", font=("微软雅黑", 10), padding=(12, 6))
        self.style.configure("Primary.TButton",
                             font=("微软雅黑", 10, "bold"),
                             padding=(14, 7))
        self.style.map("Primary.TButton",
                       background=[("active", "#1D4ED8"), ("!active", "#2563EB")],
                       foreground=[("active", "white"), ("!active", "white")])

        # 进度条
        self.style.configure("TProgressbar", thickness=12, troughcolor="#E2E8F0",
                             background="#2563EB")

        # Treeview 样式
        self.style.configure("Treeview",
                             font=("微软雅黑", 9),
                             rowheight=26,
                             background="white",
                             fieldbackground="white")
        self.style.configure("Treeview.Heading",
                             font=("微软雅黑", 9, "bold"),
                             background="#E2E8F0",
                             foreground="#1E3A5F")
        self.style.map("Treeview",
                       background=[("selected", "#DBEAFE")],
                       foreground=[("selected", "#1E3A5F")])

    # ════════════════════════════════════════════
    # 界面构建
    # ════════════════════════════════════════════

    def _build_ui(self):
        # ── 顶部操作提示 ──
        tip_frame = tk.Frame(self.root, bg="#EFF6FF", padx=12, pady=6)
        tip_frame.pack(side=tk.TOP, fill=tk.X)
        tk.Label(tip_frame,
            text="1. 点「连接」 2. 点击流股或模块 3. 通过路径或关键词找到参数 4. 添加为输入或输出 5. 配置输入变量的范围及步长 6. 开始批量模拟 |  请注意模拟完成后 Aspen 由用户手动关闭，同时建议及时保存配置",
            font=("微软雅黑", 9), fg="#2563EB", bg="#EFF6FF").pack(side=tk.LEFT)

        # ── 工具栏 ──
        toolbar = tk.Frame(self.root, bg="#FFFFFF", padx=12, pady=5)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        # 底部细线分隔
        tk.Frame(self.root, bg="#E2E8F0", height=1).pack(side=tk.TOP, fill=tk.X)

        self.connect_btn = tk.Button(
            toolbar, text="连接 Aspen Plus",
            command=self._connect_to_aspen,
            font=("微软雅黑", 10, "bold"), bg="#2563EB", fg="white",
            activebackground="#1D4ED8", activeforeground="white",
            relief=tk.FLAT, padx=16, pady=6, cursor="hand2", bd=0
        )
        self.connect_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.refresh_btn = tk.Button(
            toolbar, text="刷新变量",
            command=self._refresh_variables,
            font=("微软雅黑", 10), bg="#F1F5F9", fg="#334155",
            activebackground="#E2E8F0", activeforeground="#1E3A5F",
            relief=tk.FLAT, padx=14, pady=6, cursor="hand2", bd=0,
            state=tk.DISABLED
        )
        self.refresh_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.diag_btn = tk.Button(
            toolbar, text="诊断环境",
            command=self._run_diagnose,
            font=("微软雅黑", 10), bg="#F1F5F9", fg="#64748B",
            activebackground="#E2E8F0", activeforeground="#334155",
            relief=tk.FLAT, padx=14, pady=6, cursor="hand2", bd=0
        )
        self.diag_btn.pack(side=tk.LEFT)

        # 分隔线
        tk.Frame(toolbar, bg="#E2E8F0", width=1, height=24).pack(side=tk.LEFT, padx=(12, 12))

        self.export_btn = tk.Button(
            toolbar, text="导出配置",
            command=self._export_config,
            font=("微软雅黑", 10), bg="#F1F5F9", fg="#334155",
            activebackground="#E2E8F0", activeforeground="#1E3A5F",
            relief=tk.FLAT, padx=14, pady=6, cursor="hand2", bd=0
        )
        self.export_btn.pack(side=tk.LEFT, padx=(0, 4))

        self.import_btn = tk.Button(
            toolbar, text="导入配置",
            command=self._import_config,
            font=("微软雅黑", 10), bg="#F1F5F9", fg="#334155",
            activebackground="#E2E8F0", activeforeground="#1E3A5F",
            relief=tk.FLAT, padx=14, pady=6, cursor="hand2", bd=0
        )
        self.import_btn.pack(side=tk.LEFT)

        self.conn_label = tk.Label(toolbar, text="● 未连接",
                                    font=("微软雅黑", 10), fg="#DC2626", bg="#FFFFFF")
        self.conn_label.pack(side=tk.RIGHT, padx=(0, 4))

        # ── 主区域 ──
        main_pane = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_pane.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=(4, 0))

        # ─── 左侧面板：加载项浏览器 ───
        left_frame = ttk.LabelFrame(main_pane, text="加载项浏览器", padding=(6, 4))
        main_pane.add(left_frame, weight=1)

        # ── 顶部：流股/模块搜索（搜索框+按钮） ──
        name_search_frame = ttk.Frame(left_frame)
        name_search_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 2))
        self.name_search_var = tk.StringVar()
        self.name_search_entry = ttk.Entry(name_search_frame, textvariable=self.name_search_var)
        self.name_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.name_search_entry.bind("<Return>", lambda e: self._update_name_list(self._all_blocks, self._all_streams))
        self.name_search_btn = ttk.Button(name_search_frame, text="搜索", width=7,
                   command=lambda: self._update_name_list(self._all_blocks, self._all_streams))
        self.name_search_btn.pack(side=tk.RIGHT, padx=(4, 0))

        # ── 流股/模块名列表 ──
        name_frame = ttk.Frame(left_frame)
        name_frame.pack(side=tk.TOP, fill=tk.X, expand=False)
        self.name_tree = ttk.Treeview(name_frame, columns=("kind",), show="tree",
                                       selectmode="browse", height=8)
        self.name_tree.heading("#0", text="名称")
        self.name_tree.column("#0", width=260, minwidth=160)
        name_scroll = ttk.Scrollbar(name_frame, orient=tk.VERTICAL,
                                     command=self.name_tree.yview)
        self.name_tree.configure(yscrollcommand=name_scroll.set)
        self.name_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        name_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.name_tree.bind("<<TreeviewSelect>>", self._on_name_select)
        # 存储原始列表（搜索过滤用）
        self._all_blocks = []
        self._all_streams = []

        # ── 分割线 ──
        ttk.Separator(left_frame, orient=tk.HORIZONTAL).pack(side=tk.TOP, fill=tk.X, pady=4)

        # ── 选中提示 ──
        self.selected_label = ttk.Label(left_frame, text="选中: —",
                                         style="Status.TLabel", foreground="blue")
        self.selected_label.pack(side=tk.TOP, anchor=tk.W, pady=(0, 2))

        # ── 参数路径 ──
        path_frame = ttk.Frame(left_frame)
        path_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 2))
        ttk.Label(path_frame, text="参数路径:", font=("微软雅黑", 9)).pack(side=tk.LEFT)
        self.path_var = tk.StringVar()
        self.path_entry = ttk.Entry(path_frame, textvariable=self.path_var)
        self.path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
        self.path_entry.bind("<Return>", self._on_path_enter)
        self.path_entry.bind("<KeyRelease>", self._on_path_keyrelease)
        ttk.Button(path_frame, text="读取", width=5,
                   command=self._read_current_param).pack(side=tk.RIGHT, padx=(4, 0))
        ttk.Label(left_frame, text="说明: 输入参数路径相对于选中的模块/流股，例如 Input\\TEMP 回车读取值",
                   font=("微软雅黑", 7), foreground="gray").pack(side=tk.TOP, anchor=tk.W)

        # ── 参数搜索 ──
        search_row = ttk.Frame(left_frame)
        search_row.pack(side=tk.TOP, fill=tk.X, pady=(2, 0))
        self.param_search_var = tk.StringVar()
        self.param_search_var.trace_add("write", self._on_param_search)
        self.param_search_entry = ttk.Entry(search_row, textvariable=self.param_search_var)
        self.param_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.param_search_entry.insert(0, "搜索参数名称...")
        self.param_search_entry.bind("<FocusIn>", lambda e: self._clear_placeholder(e, self.param_search_entry, "搜索参数名称..."))
        self.param_search_entry.bind("<FocusOut>", lambda e: self._restore_placeholder(e, self.param_search_entry, "搜索参数名称..."))
        self.param_search_entry.bind("<Return>", lambda e: self._trigger_param_search())
        self.param_search_btn = ttk.Button(search_row, text="搜索", width=7,
                                           command=self._trigger_param_search)
        self.param_search_btn.pack(side=tk.RIGHT, padx=(4, 0))
        # 搜索进度条（默认隐藏）
        self.param_search_progress = ttk.Progressbar(left_frame, mode="indeterminate", length=200)

        # ── 参数搜索结果列表 ──
        param_frame = ttk.Frame(left_frame)
        param_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.param_tree = ttk.Treeview(param_frame, columns=("value", "unit", "subpath", "iotype"),
                                        show="tree headings", height=6)
        self.param_tree.heading("#0", text="名称")
        self.param_tree.heading("value", text="值")
        self.param_tree.heading("unit", text="单位")
        self.param_tree.heading("subpath", text="相对路径")
        self.param_tree.heading("iotype", text="类型")
        self.param_tree.column("#0", width=100, minwidth=60)
        self.param_tree.column("value", width=70, anchor="center")
        self.param_tree.column("unit", width=50, anchor="center")
        self.param_tree.column("subpath", width=100, minwidth=70)
        self.param_tree.column("iotype", width=50, anchor="center")
        param_scroll = ttk.Scrollbar(param_frame, orient=tk.VERTICAL,
                                      command=self.param_tree.yview)
        self.param_tree.configure(yscrollcommand=param_scroll.set)
        self.param_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        param_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.param_tree.bind("<Double-1>", self._on_param_select)

        # ── 值显示 ──
        self.val_label = tk.Label(left_frame, text="值: —",
                                   font=("微软雅黑", 11, "bold"),
                                   anchor="w", justify=tk.LEFT)
        self.val_label.pack(side=tk.TOP, anchor=tk.W, pady=(2, 0))
        self.info_label = tk.Label(left_frame, text="类型: —  单位: —",
                                    font=("微软雅黑", 8), fg="#666",
                                    anchor="w", justify=tk.LEFT)
        self.info_label.pack(side=tk.TOP, anchor=tk.W)

        # ── 添加按钮 ──
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(side=tk.TOP, fill=tk.X, pady=(4, 0))
        ttk.Button(btn_frame, text="+ 添加为输入",
                   command=self._add_custom_input).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(btn_frame, text="+ 添加为输出",
                   command=self._add_custom_output).pack(side=tk.LEFT)

        # 当前选中的路径信息
        self._selected_kind = ""   # "block" / "stream"
        self._selected_name = ""   # "B1" / "S1"
        self._cached_params = []   # 枚举缓存，供搜索使用
        self._last_full_path = ""
        self._current_full_path = ""
        self._current_val = None
        self._current_unit = ""
        self._current_dtype = ""
        self._param_tree_clicked = False
        self._param_tree_selected_path = ""

        # ─── 右侧面板 ───
        right_frame = ttk.Frame(main_pane)
        main_pane.add(right_frame, weight=2)

        notebook = ttk.Notebook(right_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # 输入变量配置标签页
        input_tab = ttk.Frame(notebook, padding=(8, 6))
        notebook.add(input_tab, text="输入变量配置")
        self._build_input_tab(input_tab)

        # 输出变量选择标签页
        output_tab = ttk.Frame(notebook, padding=(8, 6))
        notebook.add(output_tab, text="输出变量选择")
        self._build_output_tab(output_tab)

        # ─── 底部控制栏 ───
        bottom_frame = tk.Frame(self.root, bg="#FFFFFF", padx=12, pady=8)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)
        tk.Frame(self.root, bg="#E2E8F0", height=1).pack(side=tk.BOTTOM, fill=tk.X)

        self.run_btn = tk.Button(bottom_frame, text="开始批量模拟",
                                  command=self._start_batch_run,
                                  font=("微软雅黑", 11, "bold"), bg="#16A34A", fg="white",
                                  activebackground="#15803D", activeforeground="white",
                                  relief=tk.FLAT, padx=20, pady=8, cursor="hand2", bd=0,
                                  state=tk.DISABLED)
        self.run_btn.pack(side=tk.LEFT)

        self.status_label = tk.Label(bottom_frame, text="就绪 — 请先连接 Aspen Plus",
                                      font=("微软雅黑", 9), fg="#64748B", bg="#FFFFFF")
        self.status_label.pack(side=tk.LEFT, padx=(16, 0))

        self.log_toggle_btn = tk.Button(bottom_frame, text="日志",
                                         command=self._toggle_log_panel,
                                         font=("微软雅黑", 9), bg="#F1F5F9", fg="#64748B",
                                         activebackground="#E2E8F0", activeforeground="#334155",
                                         relief=tk.FLAT, padx=10, pady=4, cursor="hand2", bd=0)
        self.log_toggle_btn.pack(side=tk.RIGHT, padx=(4, 0))

        self.progress = ttk.Progressbar(bottom_frame, mode="determinate", length=280)
        self.progress.pack(side=tk.RIGHT, padx=(8, 0))
        self.progress_label = tk.Label(bottom_frame, text="0/0", width=10,
                                        font=("微软雅黑", 9), fg="#64748B", bg="#FFFFFF")
        self.progress_label.pack(side=tk.RIGHT)

        # ─── 日志面板（默认折叠） ───
        self.log_visible = False
        self.log_frame = ttk.LabelFrame(self.root, text="运行日志", padding=(6, 2))
        self.log_text = tk.Text(self.log_frame, height=8, font=("Consolas", 9),
                                 wrap=tk.WORD, state=tk.DISABLED, bg="#1e1e1e", fg="#d4d4d4")
        self.log_text.tag_configure("error", foreground="#f44747")
        self.log_text.tag_configure("warn", foreground="#cca700")
        self.log_text.tag_configure("info", foreground="#4ec9b0")
        self.log_text.tag_configure("debug", foreground="#6a9955")
        log_scroll = ttk.Scrollbar(self.log_frame, orient=tk.VERTICAL,
                                    command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 定期刷新日志显示
        self._schedule_log_refresh()

    # ──────────────────────────────────────────
    # 输入变量配置标签页
    # ──────────────────────────────────────────

    def _build_input_tab(self, parent):
        ttk.Label(parent, text="已选的输入变量（双击行可编辑范围）",
                  style="Header.TLabel").pack(anchor=tk.W, pady=(0, 6))

        columns = ("name", "path", "value", "unit", "min", "max", "step", "points", "action")
        self.input_table = ttk.Treeview(parent, columns=columns, show="headings",
                                         height=10, selectmode="extended")
        self.input_table.heading("name", text="变量名称")
        self.input_table.heading("path", text="Aspen 路径")
        self.input_table.heading("value", text="当前值")
        self.input_table.heading("unit", text="单位")
        self.input_table.heading("min", text="最小值")
        self.input_table.heading("max", text="最大值")
        self.input_table.heading("step", text="步长")
        self.input_table.heading("points", text="点数")
        self.input_table.heading("action", text="操作")

        self.input_table.column("name", width=140, minwidth=100)
        self.input_table.column("path", width=240, minwidth=160)
        self.input_table.column("value", width=80, anchor="center")
        self.input_table.column("unit", width=70, anchor="center")
        self.input_table.column("min", width=70, anchor="center")
        self.input_table.column("max", width=70, anchor="center")
        self.input_table.column("step", width=70, anchor="center")
        self.input_table.column("points", width=50, anchor="center")
        self.input_table.column("action", width=40, anchor="center")

        in_scroll_y = ttk.Scrollbar(parent, orient=tk.VERTICAL,
                                     command=self.input_table.yview)
        in_scroll_x = ttk.Scrollbar(parent, orient=tk.HORIZONTAL,
                                     command=self.input_table.xview)
        self.input_table.configure(yscrollcommand=in_scroll_y.set,
                                    xscrollcommand=in_scroll_x.set)
        self.input_table.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        in_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        in_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.input_table.bind("<Double-1>", self._edit_input_row)

        btn_frame = ttk.Frame(parent)
        btn_frame.pack(side=tk.TOP, fill=tk.X, pady=(6, 0))
        ttk.Button(btn_frame, text="删除选中", width=12,
                   command=self._delete_selected_inputs).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="清空全部", width=12,
                   command=self._clear_inputs).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="导入变量", width=12,
                   command=self._import_inputs).pack(side=tk.LEFT)
        self.combo_label = ttk.Label(btn_frame, text="预估组合数: 0",
                                      style="Status.TLabel", foreground="blue")
        self.combo_label.pack(side=tk.RIGHT)

    # ──────────────────────────────────────────
    # 输出变量选择标签页
    # ──────────────────────────────────────────

    def _build_output_tab(self, parent):
        ttk.Label(parent, text="已选的输出变量（模拟完成后将采集这些数据）",
                  style="Header.TLabel").pack(anchor=tk.W, pady=(0, 6))

        columns = ("name", "path", "value", "unit", "action")
        self.output_table = ttk.Treeview(parent, columns=columns, show="headings",
                                          height=10, selectmode="extended")
        self.output_table.heading("name", text="变量名称")
        self.output_table.heading("path", text="Aspen 路径")
        self.output_table.heading("value", text="当前值")
        self.output_table.heading("unit", text="单位")
        self.output_table.heading("action", text="操作")
        self.output_table.column("name", width=160, minwidth=100)
        self.output_table.column("path", width="360", minwidth=160)
        self.output_table.column("value", width=80, anchor="center")
        self.output_table.column("unit", width=70, anchor="center")
        self.output_table.column("action", width=40, anchor="center")

        out_scroll = ttk.Scrollbar(parent, orient=tk.VERTICAL,
                                    command=self.output_table.yview)
        self.output_table.configure(yscrollcommand=out_scroll.set)
        self.output_table.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        out_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        btn_frame = ttk.Frame(parent)
        btn_frame.pack(side=tk.TOP, fill=tk.X, pady=(6, 0))
        ttk.Button(btn_frame, text="删除选中", width=12,
                   command=self._delete_selected_outputs).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frame, text="清空全部", width=12,
                   command=self._clear_outputs).pack(side=tk.LEFT)

    # ════════════════════════════════════════════
    # 核心操作
    # ════════════════════════════════════════════

    def _connect_to_aspen(self):
        """连接到运行中的 Aspen Plus"""
        self._update_status("正在连接到 Aspen Plus...")
        self.connect_btn.config(state=tk.DISABLED, bg="#94A3B8")
        self.root.update()

        def task():
            try:
                if _HAVE_PYTHONCOM:
                    try:
                        pythoncom.CoInitialize()
                    except:
                        pass
                succ, msg = self.aspen.connect()
            except Exception as e:
                _save_crash(f"[connect线程] {type(e).__name__}: {e}\n{traceback.format_exc()}")
                succ, msg = False, f"{type(e).__name__}: {e}"
            self.root.after(0, lambda: self._on_connect_result(succ, msg))

        Thread(target=task, daemon=True).start()

    def _run_diagnose(self):
        """运行 Aspen COM 环境诊断"""
        def task():
            report = self.aspen.diagnose()
            self.root.after(0, lambda: self._show_diag_result(report))
        Thread(target=task, daemon=True).start()

    def _show_diag_result(self, report):
        dialog = tk.Toplevel(self.root)
        dialog.title("Aspen Plus 环境诊断报告")
        dialog.geometry("700x500")
        dialog.transient(self.root)
        text = tk.Text(dialog, wrap=tk.WORD, font=("Consolas", 9))
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text.insert("1.0", report)
        text.config(state=tk.DISABLED)
        btn = ttk.Button(dialog, text="关闭", command=dialog.destroy)
        btn.pack(pady=(0, 10))

    def _on_connect_result(self, succ, msg):
        self.connect_btn.config(state=tk.NORMAL, bg="#2563EB", fg="white")
        if succ:
            self.aspen_connected = True
            self.conn_label.config(text="● 已连接", fg="#16A34A")
            self._update_status(msg)

            if self.aspen.needs_file_load:
                self.root.after(500, self._prompt_load_file)
            else:
                self.refresh_btn.config(state=tk.NORMAL,
                                        bg="#F1F5F9", fg="#334155")
                self._refresh_variables()
        else:
            self.aspen_connected = False
            self.conn_label.config(text="● 未连接", fg="#DC2626")
            self.refresh_btn.config(state=tk.DISABLED,
                                    bg="#F1F5F9", fg="#94A3B8")
            self._update_status("连接失败 — 点 诊断环境 查看详情，或重试连接")
            messagebox.showerror("连接失败",
                f"{msg}\n\n"
                "请确认:\n"
                "  1. Aspen Plus 已经手动打开并加载了 .bkp 文件\n"
                "  2. 模拟已运行至收敛\n"
                "  3. 如果不确定问题，点击「诊断环境」查看详情")

    def _prompt_load_file(self):
        """Dispatch 模式：弹出文件选择对话框加载 .bkp 文件"""
        filepath = filedialog.askopenfilename(
            title="选择 Aspen Plus 模拟文件 (.bkp)",
            filetypes=[("Aspen Plus 文件", "*.bkp *.apw"), ("所有文件", "*.*")]
        )
        if not filepath:
            # 用户取消了，但仍可手动加载
            self._update_status("已连接但未加载文件 — 可通过菜单选择 .bkp 文件")
            messagebox.showinfo("提示", "已连接到 Aspen Plus，但未加载模拟文件。\n\n"
                                         "请稍后通过 Aspen Plus 窗口手动加载 .bkp 文件，\n"
                                         "然后点击「刷新变量」按钮。")
            return

        self._update_status(f"正在加载文件: {os.path.basename(filepath)}...")
        self.connect_btn.config(state=tk.DISABLED, bg="#94A3B8")
        self.root.update()

        def task():
            try:
                if _HAVE_PYTHONCOM:
                    try:
                        pythoncom.CoInitialize()
                    except:
                        pass
                succ, msg = self.aspen.load_file(filepath)
            except Exception as e:
                _save_crash(f"[load_file线程] {type(e).__name__}: {e}\n{traceback.format_exc()}")
                succ, msg = False, f"{type(e).__name__}: {e}"
            self.root.after(0, lambda: self._on_file_loaded(succ, msg))

        Thread(target=task, daemon=True).start()

    def _on_file_loaded(self, succ, msg):
        self.connect_btn.config(state=tk.NORMAL, bg="#2563EB", fg="white")
        if succ:
            self._update_status("文件加载成功，正在读取变量...")
            self.refresh_btn.config(state=tk.NORMAL, bg="#F1F5F9", fg="#334155")
            self._refresh_variables()
        else:
            self._update_status("文件加载失败")
            messagebox.showerror("文件加载失败", msg)

    def _refresh_variables(self):
        """刷新流股/模块列表"""
        if not self.aspen_connected:
            messagebox.showwarning("提示", "请先连接到 Aspen Plus")
            return

        if self.aspen.needs_file_load:
            messagebox.showwarning("提示", "请先通过文件对话框加载 .bkp 文件")
            return

        self._update_status("正在读取变量...")
        self.name_tree.delete(*self.name_tree.get_children())
        self.param_tree.delete(*self.param_tree.get_children())
        self.all_input_vars.clear()
        self.all_output_vars.clear()
        self.blocks_info.clear()
        self.streams_info.clear()
        self._cached_params = []

        def task():
            try:
                if _HAVE_PYTHONCOM:
                    try:
                        pythoncom.CoInitialize()
                    except:
                        pass
                blocks = self.aspen.get_blocks()
                streams = self.aspen.get_streams()
                self.blocks_info = blocks
                self.streams_info = streams
                self._all_blocks = blocks
                self._all_streams = streams
                self.root.after(0, lambda: self._update_name_list(blocks, streams))
            except Exception as e:
                self.root.after(0, lambda: self._update_status(f"读取变量失败: {e}"))
                _save_crash(f"[refresh_variables线程] {type(e).__name__}: {e}\n{traceback.format_exc()}")

        Thread(target=task, daemon=True).start()

    def _update_name_list(self, blocks, streams):
        """填充流股/模块名列表（支持搜索过滤）"""
        self.name_tree.delete(*self.name_tree.get_children())
        keyword = self.name_search_var.get().strip().lower()

        if blocks:
            pnode = self.name_tree.insert("", tk.END, text="模块 (Blocks)",
                                           open=True, values=("group",))
            for blk_name, blk_type in blocks:
                if keyword and keyword not in blk_name.lower():
                    continue
                self.name_tree.insert(pnode, tk.END,
                                      text=f"{blk_name}  ({blk_type})",
                                      values=("block",))

        if streams:
            pnode = self.name_tree.insert("", tk.END, text="流股 (Streams)",
                                           open=True, values=("group",))
            for sname in streams:
                if keyword and keyword not in sname.lower():
                    continue
                self.name_tree.insert(pnode, tk.END,
                                      text=sname,
                                      values=("stream",))

        self._update_run_button()
        self._update_status(
            f"就绪 — 发现 {len(blocks)} 个模块, {len(streams)} 条流股\n"
            f"在列表中点击选择，下方输入参数路径读取值"
        )

    # ──────────────────────────────────────────
    # 左侧面板交互
    # ──────────────────────────────────────────

    def _on_name_search(self, *args):
        """流股/模块搜索框输入时过滤列表"""
        self._update_name_list(self._all_blocks, self._all_streams)

    def _on_name_select(self, event):
        """点击流股/模块名时记录选中，清空之前的结果"""
        sel = self.name_tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.name_tree.item(item, "values")
        kind = vals[0] if vals else ""
        if kind in ("", "group"):
            return
        text = self.name_tree.item(item, "text")
        self._selected_kind = kind   # "block" / "stream"
        # 提取名称: "B1  (RadFrac)" → "B1", "S3" → "S3"
        name = text.strip()
        if kind == "block":
            name = name.split()[0].split("(")[0].strip()
        else:
            name = name.strip()
        self._selected_name = name
        self.selected_label.config(text=f"选中: {text}")
        # 清空旧结果
        self.param_tree.delete(*self.param_tree.get_children())
        self._cached_params = []
        self._param_tree_clicked = False
        self._param_tree_selected_path = ""
        self.val_label.config(text="值: —")
        self.info_label.config(text="类型: —  单位: —")

    def _build_full_path(self, user_path):
        """根据选中的流股/模块和用户输入的路径，构造完整 Aspen 路径"""
        if not user_path:
            return ""
        p = user_path.strip()
        # 如果用户输入了完整路径，直接使用
        if p.startswith("\\Data"):
            return p
        if not self._selected_name:
            return ""
        if self._selected_kind == "block":
            base = f"\\Data\\Blocks\\{self._selected_name}"
        elif self._selected_kind == "stream":
            base = f"\\Data\\Streams\\{self._selected_name}"
        else:
            return ""
        # 如果路径以 Input\ 或 Output\ 开头
        if p.upper().startswith("INPUT\\") or p.upper().startswith("OUTPUT\\"):
            return f"{base}\\{p}"
        # 否则默认补到 Input\ 下
        return f"{base}\\Input\\{p}"

    def _on_path_enter(self, event=None):
        """用户按回车时读取参数值"""
        self._read_current_param()

    def _on_path_keyrelease(self, event=None):
        """用户输入时稍作延迟后自动读取（避免频繁请求）"""
        if event and event.keysym in ("Return", "Up", "Down", "Left", "Right"):
            return  # 回车由 _on_path_enter 处理
        # 消抖：取消之前的定时器
        if hasattr(self, "_path_timer") and self._path_timer:
            self.root.after_cancel(self._path_timer)
        self._path_timer = self.root.after(500, self._read_current_param)

    def _read_current_param(self):
        """读取当前参数路径的值"""
        user_path = self.path_var.get().strip()
        if not user_path or not self._selected_name:
            self.val_label.config(text="值: —  (请先选择流股/模块，再输入路径)")
            self.info_label.config(text="类型: —  单位: —")
            return
        full_path = self._build_full_path(user_path)
        if not full_path:
            self.val_label.config(text="值: —  (路径构造失败)")
            return
        # 显示构造出的完整路径，方便用户验证
        self._last_full_path = full_path
        self._update_status(f"正在读取: {full_path}")

        # 检查 Aspen 是否还活着
        if not self.aspen_connected:
            self.val_label.config(text="值: —  (Aspen 连接已断开)")
            self._update_status("错误: Aspen Plus 连接已断开，请重新连接")
            return

        def task():
            try:
                val, unit, dtype = self.aspen.get_node_info(full_path)
                self.root.after(0, lambda: self._update_param_display(
                    val, unit, dtype, full_path))
            except Exception as e:
                self.root.after(0, lambda: self._update_status(
                    f"读取失败: {e}\n"
                    f"提示: 检查路径是否正确，或 Aspen Plus 是否还在运行。"
                    f"\n完整路径: {full_path}"))

        Thread(target=task, daemon=True).start()

    @staticmethod
    def _clear_placeholder(event, entry, placeholder):
        if entry.get() == placeholder:
            entry.delete(0, tk.END)

    @staticmethod
    def _restore_placeholder(event, entry, placeholder):
        if not entry.get().strip():
            entry.insert(0, placeholder)

    def _update_param_display(self, val, unit, dtype, full_path):
        if val is not None:
            self.val_label.config(text=f"值: {val}")
        else:
            self.val_label.config(text="值: (无法读取 — Aspen 中该路径无值或无此节点)")
        unit_str = unit if unit else ""
        dtype_str = dtype if dtype else "—"
        self.info_label.config(text=f"类型: {dtype_str}  单位: {unit_str}")
        self._last_full_path = full_path
        self._current_full_path = full_path
        self._current_val = val
        self._current_unit = unit_str
        self._current_dtype = dtype_str

    def _on_param_search(self, *args):
        """搜索参数列表（消抖 300ms）"""
        keyword = self.param_search_var.get().strip()
        if keyword in ("搜索参数名称...", ""):
            self.param_tree.delete(*self.param_tree.get_children())
            return
        if hasattr(self, "_search_timer") and self._search_timer:
            self.root.after_cancel(self._search_timer)
        self._search_timer = self.root.after(300, lambda: self._do_param_search(keyword))

    def _trigger_param_search(self):
        """搜索按钮 / Enter 键触发：立即搜索并显示进度条"""
        keyword = self.param_search_var.get().strip()
        if keyword in ("搜索参数名称...", ""):
            return
        # 取消消抖定时器
        if hasattr(self, "_search_timer") and self._search_timer:
            self.root.after_cancel(self._search_timer)
            self._search_timer = None
        self._do_param_search(keyword)

    def _do_param_search(self, keyword):
        """在选中的流股/模块下搜索参数，支持 \\ 路径分隔符精确搜索"""
        if not self._selected_name:
            messagebox.showinfo("提示", "请先在顶部列表中选择一个流股或模块")
            return

        # 显示进度条
        self.param_search_progress.pack(side=tk.TOP, fill=tk.X, pady=(2, 0))
        self.param_search_progress.start(10)

        kl = keyword.lower()
        # 用反斜杠分隔，支持 TEMP\MIXED 这种多层级搜索
        parts = [p.strip() for p in kl.split("\\") if p.strip()]

        def task():
            try:
                params = self.aspen.enumerate_params(
                    self._selected_kind, self._selected_name)
                self._cached_params = params
                matched = []
                for name, path, val, unit, dtype, is_num in params:
                    path_lower = path.lower()
                    # 多个层级片段全部匹配（如 TEMP\MIXED 匹配 ...\TEMP\MIXED）
                    all_match = all(p in path_lower for p in parts)
                    single_match = kl in name.lower() or kl in path_lower
                    if all_match or single_match:
                        pp = path_lower
                        # 提取 Input/Output 之后的相对路径
                        for cut in ("\\input\\", "\\output\\"):
                            idx = pp.find(cut)
                            if idx >= 0:
                                subpath = path[idx + len(cut):]
                                break
                        else:
                            subpath = path
                        # 判断输入/输出类型：路径第4层
                        path_segs = [s for s in path.split("\\") if s]
                        iotype = ""
                        if len(path_segs) >= 4:
                            seg4 = path_segs[3].lower()
                            if seg4 == "input":
                                iotype = "输入"
                            elif seg4 == "output":
                                iotype = "输出"
                        matched.append((name, path, val, unit, dtype, is_num, subpath, iotype))

                self.root.after(0, lambda: self._show_param_results(matched, kl))
            except Exception as e:
                self.root.after(0, lambda: self._update_status(f"搜索失败: {e}"))

        Thread(target=task, daemon=True).start()

    def _show_param_results(self, matched, keyword):
        # 隐藏进度条
        self.param_search_progress.stop()
        self.param_search_progress.pack_forget()
        self.param_tree.delete(*self.param_tree.get_children())
        for name, path, val, unit, dtype, is_num, subpath, iotype in matched:
            display_name = subpath if subpath else name
            val_str = f"{val}" if val is not None else "—"
            unit_str = unit if unit else ""
            self.param_tree.insert("", tk.END, text=display_name,
                                    values=(val_str, unit_str, subpath, iotype, path))
        self._update_status(f"搜索 '{keyword}': 找到 {len(matched)} 个参数")

    def _on_param_select(self, event):
        """双击搜索结果时，填入路径框并读取，记录选中项供添加使用"""
        sel = self.param_tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.param_tree.item(item, "values")
        subpath = vals[2] if vals and len(vals) > 2 else ""
        full_path = vals[4] if vals and len(vals) > 4 else ""
        self._param_tree_clicked = True
        self._param_tree_selected_path = full_path
        if subpath:
            self.path_var.set(subpath)
            self._read_current_param()

    def _add_custom_input(self):
        """将当前参数添加为输入变量
        优先使用搜索结果点击项，否则使用路径框读取项
        """
        # ── 情况1：用户点击了搜索结果中的某一项 ──
        if self._param_tree_clicked and self._param_tree_selected_path:
            sel = self.param_tree.selection()
            if sel:
                vals = self.param_tree.item(sel[0], "values")
                name = self.param_tree.item(sel[0], "text")
                full_path = vals[4] if vals and len(vals) > 4 else self._param_tree_selected_path
                val_str = vals[0] if vals else "—"
                unit_str = vals[1] if vals and len(vals) > 1 else ""
                try:
                    val = float(val_str) if val_str not in ("—", "") else None
                except:
                    val = None
                display = name
                self._confirm_add_input(display, full_path, unit_str, val)
                return

        # ── 情况2：用户通过路径框读取了值 ──
        if not self._selected_name or not self._current_full_path:
            messagebox.showinfo("提示", "请先选择流股/模块并输入参数路径读取值，或双击搜索结果项")
            return
        full_path = self._current_full_path
        if not full_path:
            return
        user_path = self.path_var.get().strip()
        display = user_path.split("\\")[-1] if user_path else full_path.split("\\")[-1]
        unit = getattr(self, "_current_unit", "")
        self._confirm_add_input(display, full_path, unit, self._current_val)

    def _confirm_add_input(self, display, full_path, unit, current_val):
        """内部：实际执行添加到输入配置"""
        for cfg in self.input_configs:
            if cfg["path"].lower() == full_path.lower():
                messagebox.showinfo("提示", "该参数已在输入列表中")
                return
        try:
            cur = float(current_val) if current_val is not None else 0.0
        except (ValueError, TypeError):
            cur = 0.0
        if cur == 0.0:
            min_val, max_val, step = -10, 10, 1
        else:
            min_val = round(cur * 0.8, 4)
            max_val = round(cur * 1.2, 4)
            step = round((max_val - min_val) / 10, 4)
            if step <= 0:
                step = 0.1
        self.input_configs.append({
            "name": display,
            "path": full_path,
            "unit": unit,
            "value": current_val,
            "min": min_val,
            "max": max_val,
            "step": step
        })
        self._refresh_input_table()
        self._update_status(f"已添加输入变量: {display}")

    def _add_custom_output(self):
        """将当前参数添加为输出变量
        优先使用搜索结果点击项，否则使用路径框读取项
        """
        # ── 情况1：用户点击了搜索结果中的某一项 ──
        if self._param_tree_clicked and self._param_tree_selected_path:
            sel = self.param_tree.selection()
            if sel:
                vals = self.param_tree.item(sel[0], "values")
                name = self.param_tree.item(sel[0], "text")
                full_path = vals[4] if vals and len(vals) > 4 else self._param_tree_selected_path
                unit_str = vals[1] if vals and len(vals) > 1 else ""
                self._confirm_add_output(name, full_path, unit_str)
                return

        # ── 情况2：用户通过路径框读取了值 ──
        if not self._selected_name or not self._current_full_path:
            messagebox.showinfo("提示", "请先选择流股/模块并输入参数路径读取值，或双击搜索结果项")
            return
        full_path = self._current_full_path
        if not full_path:
            return
        user_path = self.path_var.get().strip()
        display = user_path.split("\\")[-1] if user_path else full_path.split("\\")[-1]
        unit = getattr(self, "_current_unit", "")
        self._confirm_add_output(display, full_path, unit)

    def _confirm_add_output(self, display, full_path, unit):
        """内部：实际执行添加到输出配置。流股 Input 路径自动转为 Output 路径"""
        # 流股类型的输入路径自动转为输出路径
        out_path = AspenInterface.input_path_to_output_path(full_path)
        for cfg in self.output_configs:
            if cfg["path"].lower() == out_path.lower():
                messagebox.showinfo("提示", "该参数已在输出列表中")
                return
        current_val = None
        try:
            current_val = self.aspen.get_variable(out_path)
        except:
            pass
        self.output_configs.append({
            "name": display,
            "path": out_path,
            "unit": unit,
            "value": current_val
        })
        self._refresh_output_table()
        self._update_status(f"已添加输出变量: {display}")

    def _add_input_var(self, display_text, path, unit=""):
        """（保留）供其他模块直接调用——按路径添加输入变量"""
        path_lower = path.lower()
        for cfg in self.input_configs:
            if cfg["path"].lower() == path_lower:
                return
        current_val = self.aspen.get_variable(path)
        try:
            cur = float(current_val) if current_val is not None else 0.0
        except (ValueError, TypeError):
            cur = 0.0
        if cur == 0.0:
            min_val, max_val, step = -10, 10, 1
        else:
            min_val = round(cur * 0.8, 4)
            max_val = round(cur * 1.2, 4)
            step = round((max_val - min_val) / 10, 4)
            if step <= 0:
                step = 0.1
        name = display_text.split("=")[0].strip()
        self.input_configs.append({"name": name, "path": path, "unit": unit,
                                    "value": current_val,
                                    "min": min_val, "max": max_val, "step": step})
        self._refresh_input_table()

    def _add_output_var(self, display_text, path, unit=""):
        """（保留）供其他模块直接调用——按路径添加输出变量"""
        for cfg in self.output_configs:
            if cfg["path"].lower() == path.lower():
                return
        name = display_text.split("=")[0].strip()
        current_val = self.aspen.get_variable(path)
        self.output_configs.append({"name": name, "path": path, "unit": unit,
                                     "value": current_val})
        self._refresh_output_table()

    # ──────────────────────────────────────────
    # 输入配置表操作
    # ──────────────────────────────────────────

    def _refresh_input_table(self):
        self.input_table.delete(*self.input_table.get_children())
        total_combos = 1
        for cfg in self.input_configs:
            min_v = cfg["min"]
            max_v = cfg["max"]
            step = cfg["step"]
            unit = cfg.get("unit", "")
            val = cfg.get("value", "")
            val_str = str(val) if val is not None else "-"
            if step > 0:
                points = int((max_v - min_v) / step) + 1
            else:
                points = 1
            total_combos *= points
            self.input_table.insert("", tk.END, values=(
                cfg["name"], cfg["path"], val_str, unit,
                min_v, max_v, step, points, "✕"
            ))
        self.combo_label.config(text=f"预估组合数: {total_combos}")
        self._update_run_button()

    def _edit_input_row(self, event):
        sel = self.input_table.selection()
        if not sel:
            return
        item = sel[0]
        idx = self.input_table.index(item)
        cfg = self.input_configs[idx]

        def on_edit_cb(new_cfg):
            self._on_input_edited(idx, new_cfg)

        EditInputDialog(self.root, cfg, on_edit_cb)

    def _on_input_edited(self, idx, new_cfg):
        if idx < len(self.input_configs):
            self.input_configs[idx] = new_cfg
            self._refresh_input_table()

    def _delete_selected_inputs(self):
        sel = self.input_table.selection()
        if not sel:
            return
        indices = [self.input_table.index(item) for item in sel]
        for idx in sorted(indices, reverse=True):
            if idx < len(self.input_configs):
                self.input_configs.pop(idx)
        self._refresh_input_table()

    def _clear_inputs(self):
        if not self.input_configs:
            return
        if messagebox.askyesno("确认", "确定清空全部输入变量配置？"):
            self.input_configs.clear()
            self._refresh_input_table()

    def _import_inputs(self):
        filepath = filedialog.askopenfilename(
            title="导入输入配置",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if not filepath:
            return
        try:
            import csv
            count = 0
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    unit = row.get("unit", row.get("单位", ""))
                    value = row.get("value", row.get("当前值", ""))
                    cfg = {
                        "name": row.get("name", row.get("变量名称", "未知")),
                        "path": row.get("path", row.get("变量路径", "")),
                        "unit": unit,
                        "value": value,
                        "min": float(row.get("min", row.get("最小值", 0))),
                        "max": float(row.get("max", row.get("最大值", 10))),
                        "step": float(row.get("step", row.get("步长", 1))),
                    }
                    if cfg["path"]:
                        self.input_configs.append(cfg)
                        count += 1
            self._refresh_input_table()
            self._update_status(f"已导入 {count} 个输入变量")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    # ──────────────────────────────────────────
    # 输出配置表操作
    # ──────────────────────────────────────────

    def _refresh_output_table(self):
        self.output_table.delete(*self.output_table.get_children())
        for cfg in self.output_configs:
            unit = cfg.get("unit", "")
            val = cfg.get("value", "")
            val_str = str(val) if val is not None else "-"
            self.output_table.insert("", tk.END, values=(
                cfg["name"], cfg["path"], val_str, unit, "✕"
            ))
        self._update_run_button()

    def _delete_selected_outputs(self):
        sel = self.output_table.selection()
        if not sel:
            return
        indices = [self.output_table.index(item) for item in sel]
        for idx in sorted(indices, reverse=True):
            if idx < len(self.output_configs):
                self.output_configs.pop(idx)
        self._refresh_output_table()

    def _clear_outputs(self):
        if not self.output_configs:
            return
        if messagebox.askyesno("确认", "确定清空全部输出变量配置？"):
            self.output_configs.clear()
            self._refresh_output_table()

    # ──────────────────────────────────────────
    # 导入 / 导出配置（JSON，.log 文件）
    # ──────────────────────────────────────────

    def _export_config(self):
        """导出输入参数配置和输出变量选择，保存为 .log 文件"""
        if not self.input_configs and not self.output_configs:
            messagebox.showinfo("提示", "当前没有可导出的配置")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出配置",
            defaultextension=".log",
            filetypes=[("日志文件", "*.log"), ("所有文件", "*.*")],
            initialfile="aspen_config.log"
        )
        if not filepath:
            return

        # 构建导出数据结构
        export_data = {
            "version": 1,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "input_configs": [],
            "output_configs": []
        }
        for cfg in self.input_configs:
            export_data["input_configs"].append({
                "name": cfg["name"],
                "path": cfg["path"],
                "unit": cfg.get("unit", ""),
                "value": cfg.get("value"),
                "min": cfg["min"],
                "max": cfg["max"],
                "step": cfg["step"]
            })
        for cfg in self.output_configs:
            export_data["output_configs"].append({
                "name": cfg["name"],
                "path": cfg["path"],
                "unit": cfg.get("unit", ""),
                "value": cfg.get("value")
            })

        try:
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            self._update_status(f"配置已导出: {os.path.basename(filepath)} "
                                f"（输入 {len(self.input_configs)} 个, 输出 {len(self.output_configs)} 个）")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _import_config(self):
        """从 .log 文件导入配置，恢复输入参数配置和输出变量选择"""
        filepath = filedialog.askopenfilename(
            title="导入配置",
            filetypes=[("日志文件", "*.log"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("读取失败", f"无法解析配置文件:\n{e}")
            return

        if "input_configs" not in data and "output_configs" not in data:
            messagebox.showerror("格式错误", "配置文件不包含有效的 input_configs / output_configs 字段")
            return

        # 恢复输入配置
        in_count = 0
        old_paths = {cfg["path"].lower() for cfg in self.input_configs}
        for in_cfg in data.get("input_configs", []):
            path = in_cfg.get("path", "")
            if not path:
                continue
            if path.lower() in old_paths:
                continue  # 已存在的跳过
            self.input_configs.append({
                "name": in_cfg.get("name", path.split("\\")[-1]),
                "path": path,
                "unit": in_cfg.get("unit", ""),
                "value": in_cfg.get("value"),
                "min": in_cfg.get("min", 0),
                "max": in_cfg.get("max", 1),
                "step": in_cfg.get("step", 0.1)
            })
            old_paths.add(path.lower())
            in_count += 1
        self._refresh_input_table()

        # 恢复输出配置
        out_count = 0
        old_out_paths = {cfg["path"].lower() for cfg in self.output_configs}
        for out_cfg in data.get("output_configs", []):
            path = out_cfg.get("path", "")
            if not path:
                continue
            if path.lower() in old_out_paths:
                continue
            out_path = AspenInterface.input_path_to_output_path(path)
            self.output_configs.append({
                "name": out_cfg.get("name", path.split("\\")[-1]),
                "path": out_path,
                "unit": out_cfg.get("unit", ""),
                "value": out_cfg.get("value")
            })
            old_out_paths.add(out_path.lower())
            out_count += 1
        self._refresh_output_table()

        self._update_run_button()
        self._update_status(f"配置已导入: {os.path.basename(filepath)} "
                            f"（新增输入 {in_count} 个, 输出 {out_count} 个）")

    # ──────────────────────────────────────────
    # 运行批量模拟
    # ──────────────────────────────────────────

    def _update_run_button(self):
        has_inputs = len(self.input_configs) > 0
        has_outputs = len(self.output_configs) > 0
        enabled = has_inputs and has_outputs and self.aspen_connected
        if enabled:
            self.run_btn.config(state=tk.NORMAL, bg="#16A34A", fg="white")
        else:
            self.run_btn.config(state=tk.DISABLED, bg="#94A3B8", fg="white")

    def _start_batch_run(self):
        if not self.aspen_connected:
            messagebox.showwarning("提示", "请先连接到 Aspen Plus")
            return
        if self.aspen.needs_file_load:
            messagebox.showwarning("提示", "请先加载 .bkp 模拟文件")
            return
        if not self.input_configs:
            messagebox.showwarning("提示", "请至少配置一个输入变量")
            return
        if not self.output_configs:
            messagebox.showwarning("提示", "请至少选择一个输出变量")
            return

        # ── 选择 CSV 保存路径 ──
        csv_path = filedialog.asksaveasfilename(
            title="选择模拟结果保存路径",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
            initialfile="simulation_results.csv"
        )
        if not csv_path:
            return

        # ── 打开 CSV 文件 ──
        try:
            self._csv_fh = open(csv_path, 'w', newline='', encoding='utf-8-sig')
        except Exception as e:
            messagebox.showerror("文件错误", f"无法打开输出文件:\n{e}")
            return

        var_configs = [(cfg["path"], cfg["min"], cfg["max"], cfg["step"], cfg["name"], cfg.get("unit", ""))
                       for cfg in self.input_configs]
        output_spec = [(cfg["path"], cfg["name"], cfg.get("unit", ""))
                       for cfg in self.output_configs]

        total = 1
        for cfg in self.input_configs:
            if cfg["step"] > 0:
                points = int((cfg["max"] - cfg["min"]) / cfg["step"]) + 1
            else:
                points = 1
            total *= points

        if total > 5000:
            if not messagebox.askyesno("确认", f"共 {total} 种组合，可能需要较长时间，是否继续？"):
                self._csv_fh.close()
                self._csv_fh = None
                return

        self._csv_writer = csv.writer(self._csv_fh)
        # 写入表头: 序号 | 状态 | [输入] 名称(单位) | [输出] 名称(单位) | 备注
        header = ["序号", "状态"]
        for cfg in var_configs:
            name, unit = cfg[4], cfg[5]
            unit_tag = f" ({unit})" if unit else ""
            header.append(f"{name}{unit_tag}")
        for _p, name, unit in output_spec:
            unit_tag = f" ({unit})" if unit else ""
            header.append(f"{name}{unit_tag}")
        header.append("备注")
        self._csv_writer.writerow(header)
        self._csv_fh.flush()

        self._csv_path = csv_path
        self.simulation_results = []
        self.progress["maximum"] = total
        self.progress["value"] = 0
        self.progress_label.config(text=f"0/{total}")
        self.run_btn.config(state=tk.DISABLED, bg="#94A3B8", fg="white")
        self._update_status(f"正在模拟... (共 {total} 种组合) → {os.path.basename(csv_path)}")

        # ── Smoketest 预检：验证 COM 链路是否通畅 ──
        test_cfg = self.input_configs[0]
        test_path = test_cfg["path"]
        try:
            test_val = float(test_cfg.get("value", 0))
        except (ValueError, TypeError):
            test_val = 0.0
        _aspen_log.info(f"预检: set_variable({test_path}, {test_val})...")
        try:
            test_ok = self.aspen.set_variable(test_path, test_val)
        except Exception as e:
            _aspen_log.error(f"预检失败 (COM异常): {type(e).__name__}: {e}")
            self._close_csv()
            self.run_btn.config(state=tk.NORMAL, bg="#16A34A", fg="white")
            self._update_status("预检失败 — COM 通信异常，请重新连接 Aspen Plus")
            messagebox.showerror("预检失败",
                f"COM 通信异常:\n{type(e).__name__}: {e}\n\n"
                "请检查:\n"
                "  1. Aspen Plus 是否仍在运行\n"
                "  2. 点击「连接 Aspen Plus」重新连接后重试")
            return
        if not test_ok:
            _aspen_log.error(f"预检失败: 无法设置 {test_path} = {test_val}")
            self._close_csv()
            self.run_btn.config(state=tk.NORMAL, bg="#16A34A", fg="white")
            self._update_status("预检失败 — 节点未找到，请刷新变量后重试")
            messagebox.showerror("预检失败",
                f"无法设置变量:\n{test_path} = {test_val}\n\n"
                "请检查:\n"
                "  1. 模拟文件是否已加载\n"
                "  2. 点击「刷新变量」后重试")
            return
        _aspen_log.info("预检通过: COM 链路正常")

        def task():
            try:
                self._batch_run_worker(var_configs, output_spec, total)
            except BaseException as e:
                self.root.after(0, lambda: self._on_batch_error(e))
                # 写入 crash log
                try:
                    log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'logs')
                    os.makedirs(log_dir, exist_ok=True)
                    with open(os.path.join(log_dir, 'crash_batch.txt'), 'w', encoding='utf-8') as cf:
                        cf.write(traceback.format_exc())
                except:
                    pass

        Thread(target=task, daemon=True).start()

    def _batch_run_worker(self, var_configs, output_spec, total):
        """批量运行核心循环。模拟收敛失败→Reinit跳过；COM崩溃→重连重试；程序错误→中止"""
        paths = [cfg[0] for cfg in var_configs]
        ranges = []
        for cfg in var_configs:
            min_val, max_val, step = cfg[1], cfg[2], cfg[3]
            if step <= 0:
                values = [min_val]
            else:
                count = int((max_val - min_val) / step) + 1
                values = [round(min_val + i * step, 8) for i in range(count)]
                if values[-1] < max_val - 1e-10:
                    values.append(max_val)
            ranges.append(values)

        output_paths = [cfg[0] for cfg in output_spec]
        combo_idx = 0
        error_count = 0

        for combo in product(*ranges):
            combo_idx += 1
            input_record = {}

            # ── 第1步：设置输入变量 ──
            try:
                for path, val in zip(paths, combo):
                    ok = self.aspen.set_variable(path, val)
                    if not ok:
                        raise RuntimeError(f"设置 {path} = {val} 失败（节点未找到）")
                    input_record[path] = val
            except Exception as e:
                _aspen_log.error(f"批量运行设置变量异常: {traceback.format_exc()}")
                self._write_csv_row(combo_idx, "中止", input_record, {}, f"设置变量异常: {e}")
                self._flush_csv()
                self.root.after(0, lambda: self._on_batch_error(e))
                return

            # ── 第2步：运行模拟（含崩溃恢复） ──
            run_ok = self._run_with_crash_recovery(combo_idx, input_record, total)
            if run_ok is None:
                # 崩溃恢复失败，批次中止
                return
            if not run_ok:
                # 模拟失败（已处理）
                error_count += 1
                self.root.after(0, lambda ci=combo_idx, t=total: self._update_progress(ci, t))
                continue

            # ── 第3步：读取输出变量 ──
            output_record = {}
            read_errors = []
            for out_path, out_name, out_unit in output_spec:
                try:
                    val = self.aspen.get_variable(out_path)
                    output_record[out_path] = val
                except Exception as e:
                    output_record[out_path] = f"读取失败: {e}"
                    read_errors.append(out_name)

            if read_errors:
                self._write_csv_row(combo_idx, "部分失败", input_record, output_record,
                                    f"输出读取失败: {', '.join(read_errors)}")
            else:
                self._write_csv_row(combo_idx, "成功", input_record, output_record, "")

            self._flush_csv()
            self.root.after(0, lambda ci=combo_idx, t=total: self._update_progress(ci, t))

        # ── 完成 ──
        self.root.after(0, self._on_batch_complete, total, error_count)

    def _run_with_crash_recovery(self, combo_idx, input_record, total):
        """运行模拟，含 COM 崩溃恢复。返回 True(成功)/False(失败)/None(崩溃恢复失败需中止)"""
        success, msg, eng_status = self.aspen.run_simulation()

        if success:
            return True

        if eng_status == 4:
            # COM 服务器崩溃 → 尝试恢复
            self._write_csv_row(combo_idx, "崩溃恢复中", input_record, {}, f"COM崩溃: {msg}")
            self._flush_csv()
            recovered, rec_msg = self.aspen.reconnect_and_reload()
            if not recovered:
                self._write_csv_row(combo_idx, "中止", input_record, {},
                                    f"COM崩溃且恢复失败: {rec_msg}")
                self._flush_csv()
                self.root.after(0, lambda: self._on_batch_error(RuntimeError(rec_msg)))
                return None

            # 恢复成功 → 重新设置变量并重试
            for path, val in input_record.items():
                self.aspen.set_variable(path, val)
            success2, msg2, eng2 = self.aspen.run_simulation()
            if success2:
                return True
            # 重试仍失败（当作普通失败）
            try:
                self.aspen.reinit()
            except:
                pass
            self._write_csv_row(combo_idx, "崩溃后失败", input_record, {}, msg2)
            self._flush_csv()
            return False

        # 普通模拟失败（收敛问题等）——不调用 Reinit，保留变量状态
        self._write_csv_row(combo_idx, "失败", input_record, {}, msg)
        self._flush_csv()
        return False

    def _write_csv_row(self, idx, status, input_record, output_record, remark):
        """写入一行 CSV 数据"""
        if not hasattr(self, '_csv_writer') or self._csv_writer is None:
            return
        row = [idx, status]
        # 按输入路径顺序写入值
        for cfg in self.input_configs:
            path = cfg["path"]
            val = input_record.get(path, "")
            row.append(str(val) if val is not None else "")
        # 按输出路径顺序写入值
        for cfg in self.output_configs:
            path = cfg["path"]
            val = output_record.get(path, "")
            row.append(str(val) if val is not None else "")
        row.append(remark)
        try:
            self._csv_writer.writerow(row)
        except Exception as e:
            _save_crash(f"[_write_csv_row] CSV写入失败: {type(e).__name__}: {e}")

    def _flush_csv(self):
        """强制刷新 CSV 文件缓冲区到磁盘"""
        try:
            if hasattr(self, '_csv_fh') and self._csv_fh:
                self._csv_fh.flush()
        except:
            pass

    def _close_csv(self):
        """关闭 CSV 文件"""
        try:
            if hasattr(self, '_csv_fh') and self._csv_fh:
                self._csv_fh.close()
                self._csv_fh = None
                self._csv_writer = None
        except:
            pass

    def _update_progress(self, current, total):
        self.progress["value"] = current
        self.progress_label.config(text=f"{current}/{total}")
        basename = os.path.basename(getattr(self, '_csv_path', '')) if hasattr(self, '_csv_path') else ""
        self._update_status(f"正在模拟... ({current}/{total}) → {basename}")

    def _on_batch_complete(self, total, error_count=0):
        self._close_csv()
        self.run_btn.config(state=tk.NORMAL, bg="#16A34A", fg="white")
        self.progress["value"] = self.progress["maximum"]
        msg = f"模拟完成！共 {total} 组"
        if error_count > 0:
            msg += f"，跳过 {error_count} 组"
        self._update_status(msg)
        messagebox.showinfo("完成", msg + f"\n结果已保存至:\n{getattr(self, '_csv_path', '')}")

    def _on_batch_error(self, error):
        self._close_csv()
        self.run_btn.config(state=tk.NORMAL, bg="#16A34A", fg="white")
        if hasattr(self, '_csv_fh') and self._csv_fh:
            # 文件仍然打开但不再写入
            pass
        self._update_status(f"模拟中止: {error}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("运行错误", f"模拟已中止:\n{error}\n\n已写入的数据保存在:\n{getattr(self, '_csv_path', '')}")

    # ──────────────────────────────────────────
    # 保存到 Excel
    # ──────────────────────────────────────────

    def _save_to_excel(self):
        if not self.simulation_results:
            messagebox.showinfo("提示", "没有可保存的数据")
            return
        filepath = filedialog.asksaveasfilename(
            title="保存模拟数据",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="simulation_results.xlsx",
        )
        if not filepath:
            return
        self._update_status("正在保存到 Excel...")
        self.root.update()

        def task():
            try:
                self._export_to_excel(filepath)
                self.root.after(0, lambda: self._on_save_complete(filepath))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("保存失败", str(e)))

        Thread(target=task, daemon=True).start()

    def _export_to_excel(self, filepath):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "模拟数据"

        input_names = [cfg["name"] for cfg in self.input_configs]
        input_paths = [cfg["path"] for cfg in self.input_configs]
        output_names = [cfg["name"] for cfg in self.output_configs]
        output_paths = [cfg["path"] for cfg in self.output_configs]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        output_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = []
        for name in input_names:
            headers.append(("输入", name))
        for name in output_names:
            headers.append(("输出", name))

        for i, (category, name) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        col = 1
        for path in input_paths:
            cell = ws.cell(row=2, column=col, value=path)
            cell.font = Font(size=8, color="666666")
            cell.fill = input_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            col += 1
        for path in output_paths:
            cell = ws.cell(row=2, column=col, value=path)
            cell.font = Font(size=8, color="666666")
            cell.fill = output_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            col += 1

        for row_idx, result in enumerate(self.simulation_results, 3):
            col = 1
            for path in input_paths:
                cell = ws.cell(row=row_idx, column=col, value=result["inputs"].get(path))
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                col += 1
            for path in output_paths:
                cell = ws.cell(row=row_idx, column=col, value=result["outputs"].get(path))
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                col += 1

        for i in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=i).value or ""))
            for row in range(2, min(len(self.simulation_results) + 3, 50)):
                cell_val = str(ws.cell(row=row, column=i).value or "")
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 4, 30)

        ws.freeze_panes = "B3"

        ws2 = wb.create_sheet("配置信息")
        ws2.cell(row=1, column=1, value="参数").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="值").font = Font(bold=True)
        ws2.cell(row=2, column=1, value="模拟总组数")
        ws2.cell(row=2, column=2, value=len(self.simulation_results))
        ws2.cell(row=4, column=1, value="输入变量配置").font = Font(bold=True)
        ws2.cell(row=5, column=1, value="变量名")
        ws2.cell(row=5, column=2, value="路径")
        ws2.cell(row=5, column=3, value="最小值")
        ws2.cell(row=5, column=4, value="最大值")
        ws2.cell(row=5, column=5, value="步长")
        for i, cfg in enumerate(self.input_configs, 6):
            ws2.cell(row=i, column=1, value=cfg["name"])
            ws2.cell(row=i, column=2, value=cfg["path"])
            ws2.cell(row=i, column=3, value=cfg["min"])
            ws2.cell(row=i, column=4, value=cfg["max"])
            ws2.cell(row=i, column=5, value=cfg["step"])
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 50
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15
        ws2.column_dimensions["E"].width = 15

        wb.save(filepath)

    def _on_save_complete(self, filepath):
        self._update_status(f"数据已保存到: {filepath}")
        messagebox.showinfo("保存成功", f"数据已保存到:\n{filepath}")

    # ──────────────────────────────────────────
    # 工具方法
    # ──────────────────────────────────────────

    def _update_status(self, msg):
        self.status_label.config(text=msg)

    # ──────────────────────────────────────────
    # 日志面板
    # ──────────────────────────────────────────

    def _toggle_log_panel(self):
        if self.log_visible:
            self.log_frame.pack_forget()
            self.log_visible = False
        else:
            self.log_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, padx=10, pady=(0, 4))
            self.log_visible = True
            self._refresh_log_display()

    def _refresh_log_display(self):
        if not self.log_visible:
            return
        if _aspen_log is None:
            return
        try:
            logs = _aspen_log.get_logs(300)
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete("1.0", tk.END)
            for level, msg in logs:
                tag = level.lower()
                line = f"[{level}] {msg}\n"
                self.log_text.insert(tk.END, line, tag)
            self.log_text.config(state=tk.DISABLED)
            self.log_text.see(tk.END)
        except:
            pass

    def _schedule_log_refresh(self):
        """每 2 秒刷新日志面板"""
        if self.log_visible:
            self._refresh_log_display()
        self.root.after(2000, self._schedule_log_refresh)

    def _on_close(self):
        """关闭程序 — 仅释放 COM 引用，Aspen Plus 由用户手动管理"""
        if self.aspen_connected:
            self.aspen.close()
        self.root.destroy()


# ════════════════════════════════════════════════════════════
# 编辑对话框
# ════════════════════════════════════════════════════════════

class EditInputDialog:
    """输入变量编辑对话框（带滚动条和鼠标滚轮支持）"""

    def __init__(self, parent, config, callback):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("编辑输入变量")
        self.dialog.geometry("520x420")
        self.dialog.resizable(True, True)
        self.dialog.minsize(400, 300)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        self.config = config.copy()
        self.callback = callback

        outer = ttk.Frame(self.dialog)
        outer.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0)
        bar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        self._sf = ttk.Frame(canvas, padding=20)
        self._sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        cw = canvas.create_window((0, 0), window=self._sf, anchor="nw")
        canvas.configure(yscrollcommand=bar.set)
        canvas.bind("<Configure>", lambda ev: canvas.itemconfig(cw, width=ev.width))
        canvas.bind("<Enter>", lambda ev: canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*e.delta/120), "units")))
        canvas.bind("<Leave>", lambda ev: canvas.unbind_all("<MouseWheel>"))
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        bar.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 按钮固定在底部（不随滚动移动） ──
        bf = ttk.Frame(self.dialog, padding=(0, 8, 0, 8))
        bf.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Button(bf, text="确定", width=10, command=self._on_ok).pack(side=tk.RIGHT, padx=(4, 20))
        ttk.Button(bf, text="取消", width=10, command=self.dialog.destroy).pack(side=tk.RIGHT, padx=(0, 4))

        # 用 outer 填充剩余空间
        outer.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        f = self._sf
        config = self.config

        ttk.Label(f, text="变量名称:").grid(row=0, column=0, sticky=tk.W, pady=6)
        ttk.Label(f, text=config["name"], font=("微软雅黑", 9, "bold")).grid(
            row=0, column=1, columnspan=3, sticky=tk.W, pady=6)

        ttk.Label(f, text="Aspen 路径:").grid(row=1, column=0, sticky=tk.W, pady=6)
        ttk.Label(f, text=config["path"], wraplength=320,
                   font=("Consolas", 8), foreground="#555").grid(
            row=1, column=1, columnspan=3, sticky=tk.W, pady=6)

        ttk.Label(f, text="单位:").grid(row=2, column=0, sticky=tk.W, pady=6)
        unit = config.get("unit", "")
        common_units = ["C", "F", "K", "atm", "bar", "Pa", "kPa", "MPa",
                        "psi", "kg/hr", "kg/s", "lb/hr", "kmol/hr", "mol/s",
                        "m3/hr", "L/min", "J/hr", "kJ/hr", "kW", "MW", ""]
        self.unit_var = tk.StringVar(value=unit)
        ttk.Combobox(f, textvariable=self.unit_var, values=common_units,
                      width=12, state="normal").grid(row=2, column=1, sticky=tk.W, pady=6)

        ttk.Label(f, text="最小值:").grid(row=3, column=0, sticky=tk.W, pady=6)
        self.min_var = tk.StringVar(value=str(config["min"]))
        ttk.Entry(f, textvariable=self.min_var, width=15).grid(row=3, column=1, sticky=tk.W, pady=6)

        ttk.Label(f, text="最大值:").grid(row=4, column=0, sticky=tk.W, pady=6)
        self.max_var = tk.StringVar(value=str(config["max"]))
        ttk.Entry(f, textvariable=self.max_var, width=15).grid(row=4, column=1, sticky=tk.W, pady=6)

        ttk.Label(f, text="步长:").grid(row=5, column=0, sticky=tk.W, pady=6)
        self.step_var = tk.StringVar(value=str(config["step"]))
        ttk.Entry(f, textvariable=self.step_var, width=15).grid(row=5, column=1, sticky=tk.W, pady=6)

        self.points_label = ttk.Label(f, text="")
        self.points_label.grid(row=4, column=2, sticky=tk.W, padx=(10, 0))
        self.min_var.trace_add("write", self._update_points)
        self.max_var.trace_add("write", self._update_points)
        self.step_var.trace_add("write", self._update_points)
        self._update_points()

        self.dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 420) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 300) // 2
        self.dialog.geometry(f"+{x}+{y}")

    def _update_points(self, *args):
        try:
            min_v = float(self.min_var.get())
            max_v = float(self.max_var.get())
            step = float(self.step_var.get())
            if step > 0:
                points = int((max_v - min_v) / step) + 1
                self.points_label.config(text=f"≈ {points} 个值")
            else:
                self.points_label.config(text="步长必须 > 0")
        except:
            self.points_label.config(text="")

    def _on_ok(self):
        try:
            min_v = float(self.min_var.get())
            max_v = float(self.max_var.get())
            step = float(self.step_var.get())
            if step <= 0:
                messagebox.showerror("错误", "步长必须大于 0")
                return
            if max_v <= min_v:
                messagebox.showerror("错误", "最大值必须大于最小值")
                return
            self.config["min"] = min_v
            self.config["max"] = max_v
            self.config["step"] = step
            self.config["unit"] = self.unit_var.get()
            self.callback(self.config)
            self.dialog.destroy()
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数值")


# ════════════════════════════════════════════════════════════
# 入口
# ════════════════════════════════════════════════════════════

def main():
    """启动图形界面，捕获启动异常并写入日志文件"""
    try:
        root = tk.Tk()
        app = AspenUI(root)
        root.mainloop()
    except Exception as e:
        import traceback
        error_log = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "启动错误.log"
        )
        with open(error_log, "w", encoding="utf-8") as f:
            f.write(f"启动时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"错误信息: {e}\n")
            f.write("详细堆栈:\n")
            traceback.print_exc(file=f)
        # 尝试弹窗提示
        try:
            import tkinter.messagebox as mb
            mb.showerror("启动错误", f"程序启动失败，详情请查看启动错误.log\n\n{e}")
        except:
            pass
        raise


if __name__ == "__main__":
    main()
